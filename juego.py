import random
import datetime
import os
import getpass
import platform
import time
from openpyxl import Workbook, load_workbook

# --- COLORES ---
try:
    from termcolor import colored
    COLORES_HABILITADOS = True
except ImportError:
    COLORES_HABILITADOS = False
    def colored(text, color=None, on_color=None, attrs=None):
        return text  # Fallback sin color

# --- SISTEMA DE SONIDOS ---
class SistemaSonidos:
    """Sistema de sonidos multiplataforma simple."""
    
    def __init__(self, activado=True):
        self.activado = activado
        self.sistema = platform.system()
        self.volumen = 0.7  # Volumen por defecto (0.0 a 1.0)
        
        # Intenta importar winsound solo en Windows
        if self.sistema == "Windows":
            try:
                import winsound
                self.winsound = winsound
                self.sonidos_disponibles = True
            except ImportError:
                self.sonidos_disponibles = False
        else:
            self.sonidos_disponibles = True  # Para otros sistemas usamos métodos alternativos
        
        # Configurar colores para mensajes de sonido
        self.estado_color = 'green' if activado else 'yellow'
    
    def emitir_beep(self, frecuencia, duracion):
        """Emite un beep básico."""
        if not self.activado or not self.sonidos_disponibles:
            return
            
        try:
            if self.sistema == "Windows":
                # Ajustar frecuencia según volumen
                freq_ajustada = int(frecuencia * self.volumen)
                self.winsound.Beep(freq_ajustada, duracion)
            elif self.sistema == "Darwin":  # macOS
                # Usar el sonido del sistema
                os.system(f"osascript -e 'beep {frecuencia}'")
            else:  # Linux/Unix
                # Carácter de alerta
                print('\a', end='', flush=True)
        except:
            pass  # Silenciar errores
    
    def sonido_menu(self):
        """Sonido al navegar por menús."""
        self.emitir_beep(600, 100)
    
    def sonido_acierto(self):
        """Sonido cuando se acierta el número."""
        if not self.activado:
            return
            
        # Secuencia ascendente de tonos (victoria)
        try:
            if self.sistema == "Windows":
                tonos = [(523, 200), (659, 200), (784, 200), (1047, 300)]  # Do, Mi, Sol, Do
                for freq, dur in tonos:
                    self.emitir_beep(freq, dur)
                    time.sleep(0.05)
            else:
                # Para otros sistemas
                self.emitir_beep(800, 150)
                time.sleep(0.1)
                self.emitir_beep(1000, 150)
                time.sleep(0.1)
                self.emitir_beep(1200, 250)
        except:
            pass
    
    def sonido_error(self):
        """Sonido cuando se falla."""
        self.emitir_beep(300, 300)
    
    def sonido_victoria(self):
        """Sonido especial de victoria."""
        if not self.activado:
            return
            
        try:
            # Fanfarria de victoria
            if self.sistema == "Windows":
                secuencia = [
                    (523, 150), (0, 50), (659, 150), (0, 50),
                    (784, 150), (0, 50), (1047, 300)
                ]
                for freq, dur in secuencia:
                    if freq > 0:
                        self.emitir_beep(freq, dur)
                    else:
                        time.sleep(dur / 1000)
            else:
                for _ in range(3):
                    self.emitir_beep(800, 100)
                    time.sleep(0.05)
                self.emitir_beep(1200, 400)
        except:
            pass
    
    def sonido_derrota(self):
        """Sonido de derrota."""
        if not self.activado:
            return
            
        # Secuencia descendente triste
        for freq in [400, 350, 300, 250]:
            self.emitir_beep(freq, 150)
            time.sleep(0.05)
    
    def sonido_sorpresa(self):
        """Sonido cuando estás muy cerca."""
        self.emitir_beep(1000, 100)
        time.sleep(0.05)
        self.emitir_beep(1100, 100)
    
    def sonido_entrada(self):
        """Sonido al entrar al juego."""
        self.emitir_beep(800, 200)
        time.sleep(0.1)
        self.emitir_beep(1000, 200)
    
    def sonido_salida(self):
        """Sonido al salir del juego."""
        for freq in [1000, 800, 600]:
            self.emitir_beep(freq, 150)
            time.sleep(0.05)
    
    def toggle_activado(self):
        """Activa/desactiva los sonidos."""
        self.activado = not self.activado
        estado = "activados" if self.activado else "desactivados"
        print(colored(f"\n✓ Sonidos {estado}", self.estado_color))
        
        # Sonido de confirmación
        if self.activado:
            self.sonido_menu()
        return self.activado
    
    def ajustar_volumen(self, nuevo_volumen):
        """Ajusta el volumen (0.0 a 1.0)."""
        if 0.0 <= nuevo_volumen <= 1.0:
            self.volumen = nuevo_volumen
            print(colored(f"✓ Volumen ajustado a {nuevo_volumen:.1f}", 'green'))
            # Sonido de prueba
            if self.activado:
                self.emitir_beep(800, 100)
        else:
            print(colored("✗ El volumen debe estar entre 0.0 y 1.0", 'red'))
    
    def probar_sonidos(self):
        """Reproduce todos los sonidos para probarlos."""
        if not self.activado:
            print(colored("Los sonidos están desactivados. Actívalos primero.", 'yellow'))
            return
            
        print(colored("\n🔊 Probando sonidos...", 'cyan'))
        
        sonidos_prueba = [
            ("Menú", self.sonido_menu),
            ("Error", self.sonido_error),
            ("Cerca", self.sonido_sorpresa),
            ("Acierto", self.sonido_acierto),
            ("Victoria", self.sonido_victoria),
            ("Derrota", self.sonido_derrota)
        ]
        
        for nombre, funcion in sonidos_prueba:
            print(colored(f"  Reproduciendo: {nombre}...", 'blue'))
            funcion()
            time.sleep(0.5)
        
        print(colored("✓ Prueba completada", 'green'))

# Crear instancia global del sistema de sonidos
sonidos = SistemaSonidos(activado=True)

# --- LIMPIEZA DE PANTALLA ---
def limpiar_pantalla():
    """Limpia la pantalla de forma multiplataforma."""
    os.system('cls' if os.name == 'nt' else 'clear')

# --- ARCHIVO EXCEL ---
ARCHIVO_EXCEL = "estadisticas_adivinanza.xlsx"

def inicializar_excel():
    """Crea el archivo Excel si no existe y añade encabezados."""
    if not os.path.exists(ARCHIVO_EXCEL):
        wb = Workbook()
        ws = wb.active
        ws.title = "Estadísticas"
        ws.append(["Fecha", "Modo", "Jugador1", "Jugador2", "Dificultad",
                   "Número Secreto", "Intentos Usados", "Resultado", "Nota"])
        wb.save(ARCHIVO_EXCEL)

def guardar_partida(modo, jugador1, jugador2, dificultad, numero_secreto,
                    intentos_usados, max_intentos, ganado):
    """Guarda la partida en Excel."""
    fecha = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    resultado = "Ganado" if ganado else "Perdido"
    
    if ganado:
        nota = round((max_intentos - intentos_usados + 1) / max_intentos * 10, 2)
    else:
        nota = 0.0
    
    wb = load_workbook(ARCHIVO_EXCEL)
    ws = wb["Estadísticas"]
    ws.append([fecha, modo, jugador1, jugador2 or "", dificultad,
               numero_secreto, intentos_usados, resultado, nota])
    wb.save(ARCHIVO_EXCEL)

def mostrar_estadisticas(filtro_usuario=None):
    """Muestra estadísticas en consola con promedio y mejor nota."""
    if not os.path.exists(ARCHIVO_EXCEL):
        print(colored("No hay estadísticas guardadas aún.", 'yellow'))
        return
    
    wb = load_workbook(ARCHIVO_EXCEL)
    ws = wb["Estadísticas"]
    filas = list(ws.iter_rows(min_row=2, values_only=True))
    
    if filtro_usuario:
        filas = [f for f in filas if filtro_usuario.lower() in (f[2] or "").lower() or
                 filtro_usuario.lower() in (f[3] or "").lower()]
        print(colored(f"\n=== ESTADÍSTICAS FILTRADAS POR: {filtro_usuario.upper()} ===", 'cyan'))
    else:
        print(colored("\n=== ESTADÍSTICAS GENERALES ===", 'cyan'))
    
    if not filas:
        print(colored("No se encontraron partidas.", 'yellow'))
        return
    
    print(f"{'Fecha':20} {'Modo':10} {'Jugador1':12} {'Jugador2':12} "
          f"{'Dif.':6} {'Nº Secreto':12} {'Intentos':8} {'Resultado':10} {'Nota':8}")
    print("-" * 105)
    
    for row in filas:
        fecha, modo, j1, j2, dif, num, intentos, res, nota = row
        j2 = j2 or "-"
        num_mostrar = "***" if res == "Ganado" else num
        res_color = 'green' if res == "Ganado" else 'red'
        print(f"{fecha:20} {modo:10} {j1:12} {j2:12} {dif:6} "
              f"{num_mostrar:12} {intentos:8} {colored(res, res_color):10} {nota:.2f}")
    
    # Estadísticas avanzadas
    notas_ganadas = [row[8] for row in filas if row[7] == "Ganado" and row[8] > 0]
    if notas_ganadas:
        promedio = sum(notas_ganadas) / len(notas_ganadas)
        print(colored(f"\nPromedio de nota en partidas ganadas: {promedio:.2f}", 'magenta'))
    
    mejor = max(filas, key=lambda x: x[8] if isinstance(x[8], (int, float)) else 0)
    print(colored(f"Mejor nota: {mejor[8]:.2f} → {mejor[2]} ({mejor[0]})", 'magenta'))

def elegir_dificultad():
    """Submenú común de dificultad usado en ambos modos."""
    sonidos.sonido_menu()
    while True:
        print(colored("\nElige la dificultad:", 'blue'))
        print("1. Fácil (20 intentos)")
        print("2. Medio (12 intentos)")
        print("3. Difícil (5 intentos)")
        opc = input(colored("Opción (1-3): ", 'blue')).strip()
        
        sonidos.sonido_menu()  # Sonido al seleccionar
        
        if opc == "1":
            return "Fácil", 20
        elif opc == "2":
            return "Medio", 12
        elif opc == "3":
            return "Difícil", 5
        else:
            print(colored("Opción inválida. Intenta de nuevo.", 'red'))
            sonidos.sonido_error()

def jugar_solitario():
    sonidos.sonido_menu()  # Sonido al entrar
    
    dificultad_texto, max_intentos = elegir_dificultad()
    
    while True:
        nombre = input(colored("\nIngrese su nombre: ", 'blue')).strip()
        if len(nombre) < 2:
            print(colored("El nombre debe tener al menos 2 caracteres.", 'red'))
            sonidos.sonido_error()
        else:
            sonidos.sonido_menu()
            break
    
    numero_secreto = random.randint(1, 1000)
    intentos = 0
    
    # Sonido de inicio de partida
    sonidos.emitir_beep(800, 100)
    time.sleep(0.1)
    sonidos.emitir_beep(1000, 100)
    
    print(colored(f"\n¡Bienvenido {nombre}! Adivina el número entre 1 y 1000.", 'green'))
    print(colored(f"Dificultad {dificultad_texto} - Tienes {max_intentos} intentos.\n", 'green'))
    
    while intentos < max_intentos:
        try:
            adivinanza = int(input(colored("Ingresa tu número: ", 'blue')))
            sonidos.sonido_menu()  # Sonido al introducir número
            
            if not 1 <= adivinanza <= 1000:
                print(colored("El número debe estar entre 1 y 1000.", 'red'))
                sonidos.sonido_error()
                continue
        except ValueError:
            print(colored("Por favor, ingresa un número válido.", 'red'))
            sonidos.sonido_error()
            continue
        
        intentos += 1
        
        # Verificar si está muy cerca (diferencia <= 10)
        if abs(adivinanza - numero_secreto) <= 10 and adivinanza != numero_secreto:
            print(colored("¡Estás muy cerca! 🔥", 'cyan', attrs=['bold']))
            sonidos.sonido_sorpresa()
        
        if adivinanza == numero_secreto:
            sonidos.sonido_victoria()  # Sonido especial de victoria
            time.sleep(0.5)  # Pausa dramática
            sonidos.sonido_acierto()  # Sonido adicional
            
            print(colored(f"\n{'='*60}", 'green'))
            print(colored(f"¡FELICIDADES {nombre.upper()}! ¡Has ganado en {intentos} intentos!", 'green', attrs=['bold']))
            print(colored(f"{'='*60}", 'green'))
            
            guardar_partida("Solitario", nombre, None, dificultad_texto,
                            numero_secreto, intentos, max_intentos, True)
            return
        elif adivinanza < numero_secreto:
            print(colored("El número buscado es mayor ↑", 'yellow'))
            sonidos.sonido_error()
        else:
            print(colored("El número buscado es menor ↓", 'yellow'))
            sonidos.sonido_error()
        
        print(colored(f"Te quedan {max_intentos - intentos} intentos.\n", 'cyan'))
    
    # Si llega aquí, ha perdido
    sonidos.sonido_derrota()
    time.sleep(0.3)
    
    print(colored(f"\n{'='*60}", 'red'))
    print(colored(f"¡Has perdido! El número era {numero_secreto}.", 'red', attrs=['bold']))
    print(colored(f"{'='*60}", 'red'))
    
    guardar_partida("Solitario", nombre, None, dificultad_texto,
                    numero_secreto, intentos, max_intentos, False)

def jugar_dos_jugadores():
    sonidos.sonido_menu()  # Sonido al entrar
    
    print(colored("=== MODO 2 JUGADORES ===", 'magenta', attrs=['bold']))
    
    while True:
        jugador1 = input(colored("Nombre del Jugador 1 (quien piensa el número): ", 'blue')).strip()
        if len(jugador1) < 2:
            print(colored("Nombre demasiado corto.", 'red'))
            sonidos.sonido_error()
        else:
            sonidos.sonido_menu()
            break

    limpiar_pantalla()
    print(colored(f"{jugador1}, ahora es tu turno de introducir el número secreto.", 'yellow'))
    input(colored("Presiona Enter cuando el Jugador 2 no esté mirando...", 'yellow'))
    
    sonidos.sonido_menu()

    try:
        numero_secreto_str = getpass.getpass(prompt=colored(f"{jugador1}, ingresa el número secreto (1-1000): ", 'blue'))
        numero_secreto = int(numero_secreto_str)
        if not 1 <= numero_secreto <= 1000:
            raise ValueError
    except:
        print(colored("\nNúmero inválido. Partida cancelada.", 'red'))
        sonidos.sonido_error()
        return

    sonidos.sonido_acierto()  # Sonido de confirmación
    print(colored(f"\n✓ Número secreto registrado.", 'green'))
    
    print(colored("\n💡 Sugerencias según el número:", 'cyan'))
    if numero_secreto <= 100 or numero_secreto >= 900:
        print(colored("- Está en un extremo → más difícil de adivinar.", 'red'))
    elif numero_secreto <= 300 or numero_secreto >= 700:
        print(colored("- Algo alejado del centro → dificultad media recomendada.", 'yellow'))
    else:
        print(colored("- Cerca del centro → más fácil de adivinar.", 'green'))

    dificultad_texto, max_intentos = elegir_dificultad()

    limpiar_pantalla()
    while True:
        jugador2 = input(colored("Nombre del Jugador 2 (quien adivina): ", 'blue')).strip()
        if len(jugador2) < 2:
            print(colored("Nombre demasiado corto.", 'red'))
            sonidos.sonido_error()
        else:
            sonidos.sonido_menu()
            break

    limpiar_pantalla()
    
    # Sonido de inicio de partida
    sonidos.emitir_beep(800, 100)
    time.sleep(0.1)
    sonidos.emitir_beep(1000, 100)
    
    print(colored(f"¡Turno de {jugador2}!", 'green', attrs=['bold']))
    print(colored(f"Adivina el número que pensó {jugador1} (entre 1 y 1000).", 'green'))
    print(colored(f"Dificultad: {dificultad_texto} - Tienes {max_intentos} intentos.\n", 'green'))

    intentos = 0
    while intentos < max_intentos:
        try:
            adivinanza = int(input(colored("Ingresa tu adivinanza: ", 'blue')))
            sonidos.sonido_menu()
            
            if not 1 <= adivinanza <= 1000:
                print(colored("El número debe estar entre 1 y 1000.", 'red'))
                sonidos.sonido_error()
                continue
        except ValueError:
            print(colored("Por favor, ingresa un número válido.", 'red'))
            sonidos.sonido_error()
            continue
        
        intentos += 1
        
        # Verificar si está muy cerca
        if abs(adivinanza - numero_secreto) <= 10 and adivinanza != numero_secreto:
            print(colored("¡Estás muy cerca! 🔥", 'cyan', attrs=['bold']))
            sonidos.sonido_sorpresa()
        
        if adivinanza == numero_secreto:
            sonidos.sonido_victoria()
            time.sleep(0.5)
            sonidos.sonido_acierto()
            
            print(colored(f"\n{'='*60}", 'green'))
            print(colored(f"¡ENHORABUENA {jugador2.upper()}! ¡Has acertado en {intentos} intentos!", 'green', attrs=['bold']))
            print(colored(f"{'='*60}", 'green'))
            
            guardar_partida("2 Jugadores", jugador1, jugador2, dificultad_texto,
                            numero_secreto, intentos, max_intentos, True)
            return
        elif adivinanza < numero_secreto:
            print(colored("El número es mayor ↑", 'yellow'))
            sonidos.sonido_error()
        else:
            print(colored("El número es menor ↓", 'yellow'))
            sonidos.sonido_error()
        
        print(colored(f"Te quedan {max_intentos - intentos} intentos.\n", 'cyan'))
    
    # Derrota
    sonidos.sonido_derrota()
    time.sleep(0.3)
    
    print(colored(f"\n{'='*60}", 'red'))
    print(colored(f"¡Se acabaron los intentos! El número era {numero_secreto}.", 'red', attrs=['bold']))
    print(colored(f"{'='*60}", 'red'))
    
    guardar_partida("2 Jugadores", jugador1, jugador2, dificultad_texto,
                    numero_secreto, intentos, max_intentos, False)

def menu_estadisticas():
    sonidos.sonido_menu()
    
    while True:
        print(colored("\n--- Menú Estadísticas ---", 'cyan'))
        print("1. Ver todas las estadísticas")
        print("2. Filtrar por nombre de jugador")
        print("3. Volver al menú principal")
        opc = input(colored("Elige opción (1-3): ", 'blue')).strip()
        
        sonidos.sonido_menu()
        
        if opc == "1":
            mostrar_estadisticas()
        elif opc == "2":
            nombre = input(colored("Ingresa el nombre del jugador a filtrar: ", 'blue')).strip()
            if len(nombre) >= 2:
                sonidos.sonido_menu()
                mostrar_estadisticas(nombre)
            else:
                print(colored("Nombre no válido.", 'red'))
                sonidos.sonido_error()
        elif opc == "3":
            break
        else:
            print(colored("Opción inválida.", 'red'))
            sonidos.sonido_error()

def menu_configuracion():
    """Menú para configurar sonidos y otros ajustes."""
    global sonidos
    
    while True:
        limpiar_pantalla()
        sonidos.sonido_menu()
        
        print(colored("\n=== CONFIGURACIÓN ===", 'cyan', attrs=['bold']))
        print(f"1. {'🔇 DESACTIVAR' if sonidos.activado else '🔊 ACTIVAR'} sonidos")
        print("2. Ajustar volumen 🔊")
        print("3. Probar todos los sonidos 🎵")
        print("4. Volver al menú principal")
        
        opcion = input(colored("\nOpción (1-4): ", 'blue')).strip()
        
        sonidos.sonido_menu()
        
        if opcion == "1":
            sonidos.toggle_activado()
            input(colored("\nPresiona Enter para continuar...", 'blue'))
            
        elif opcion == "2":
            try:
                vol = float(input("Volumen (0.0 a 1.0): "))
                sonidos.ajustar_volumen(vol)
                input(colored("\nPresiona Enter para continuar...", 'blue'))
            except ValueError:
                print(colored("Por favor, ingresa un número válido.", 'red'))
                sonidos.sonido_error()
                input(colored("\nPresiona Enter para continuar...", 'blue'))
                
        elif opcion == "3":
            sonidos.probar_sonidos()
            input(colored("\nPresiona Enter para continuar...", 'blue'))
            
        elif opcion == "4":
            break
            
        else:
            print(colored("Opción inválida.", 'red'))
            sonidos.sonido_error()

def menu_principal():
    inicializar_excel()
    sonidos.sonido_entrada()  # Sonido de bienvenida
    
    print(colored("\n" + "="*60, 'magenta'))
    print(colored("         JUEGO DE ADIVINANZA - ¡CON SONIDOS! 🎵", 'magenta', attrs=['bold']))
    print(colored("="*60, 'magenta'))
    
    # Mostrar estado de sonidos
    estado_sonido = "🔊 ACTIVADOS" if sonidos.activado else "🔇 DESACTIVADOS"
    print(colored(f"\nEstado de sonidos: {estado_sonido}", 'cyan'))
    
    while True:
        print(colored("\n1. Partida modo solitario 🎮", 'white'))
        print(colored("2. Partida 2 Jugadores 👥", 'white'))
        print(colored("3. Estadísticas 📊", 'white'))
        print(colored("4. Configuración ⚙️", 'white'))
        print(colored("5. Salir 🚪", 'white'))
        
        opcion = input(colored("\nElige una opción (1-5): ", 'blue')).strip()
        
        sonidos.sonido_menu()
        
        if opcion == "1":
            jugar_solitario()
        elif opcion == "2":
            jugar_dos_jugadores()
        elif opcion == "3":
            menu_estadisticas()
        elif opcion == "4":
            menu_configuracion()
        elif opcion == "5":
            sonidos.sonido_salida()  # Sonido de despedida
            print(colored("\n" + "="*60, 'green'))
            print(colored("¡Gracias por jugar! Hasta pronto. 👋", 'green', attrs=['bold']))
            print(colored("="*60, 'green'))
            break
        else:
            print(colored("Opción no válida. Intenta de nuevo.", 'red'))
            sonidos.sonido_error()

if __name__ == "__main__":
    try:
        menu_principal()
    except KeyboardInterrupt:
        print(colored("\n\nJuego interrumpido. ¡Hasta la próxima!", 'yellow'))
        sonidos.sonido_salida()